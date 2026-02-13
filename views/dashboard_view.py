import flet as ft
from database import get_user, reset_user_progress_for_level, get_difficult_words

def DashboardView(page: ft.Page):
    import os
    import datetime
    log_file = os.path.join(os.path.expanduser("~"), "english_mastery_debug.log")
    def log(msg):
        try:
            timestamp = datetime.datetime.now().strftime("%H:%M:%S")
            with open(log_file, "a", encoding="utf-8") as f:
                f.write(f"[{timestamp}] [DASHBOARD] {msg}\n")
        except: pass

    from session_utils import get_session
    username = get_session(page, "username")
    log(f"Retrieved username: {username}")
    
    user = get_user(username)
    log(f"Database user found: {user is not None}")
    
    if not user:
        log("No user found, REDIRECTING to /")
        page.go("/")
        return ft.View(route="/dashboard-empty")

    def start_level(e, level):
        # Reset progress for this level to start from beginning
        # OR just set current level and keep index if "Continue" is desired for that specific level
        # Requirement: "Start" starts from beginning, "Continue" continues.
        # The dashboard usually just lets you pick a level.
        # Let's add a dialog or simple logic: Clicking a level card sets the level.
        # If they want to "Start Over", we might need a specific button for that, 
        # but for now, clicking a level sets it as current and goes to learning.
        # We will check if they are already in this level to decide if we resume or start over?
        # Actually requirement says: "Choice to Start from beginning ... and Choice to Continue"
        
        # Let's show a dialog to choose "Start Over" or "Continue"
        def close_dialog(e):
            # dlg.open = False
            page.close(dlg)
            page.update()

        def go_continue(e):
            # Just set the level and navigate - don't reset anything
            from session_utils import set_session
            set_session(page, "current_level", level)
            page.close(dlg)
            page.go("/learn")

        def go_start_over(e):
            reset_user_progress_for_level(username, level)
            from session_utils import set_session
            set_session(page, "current_level", level)
            page.close(dlg)
            page.go("/learn")

        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text(f"Level {level}"),
            content=ft.Text("Do you want to continue where you left off or start over?"),
            actions=[
                ft.TextButton("Continue", on_click=go_continue),
                ft.TextButton("Start Over", on_click=go_start_over),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        # page.dialog = dlg
        # dlg.open = True
        page.open(dlg)
        page.update()

    def level_card(level_num):
        return ft.Container(
            content=ft.Column(
                [
                    ft.Icon(ft.Icons.BOOK, size=40, color="white"),
                    ft.Text(f"Level {level_num}", size=20, weight=ft.FontWeight.BOLD, color="white"),
                    ft.Text("1000 Words", size=12, color="white70"),
                ],
                alignment=ft.MainAxisAlignment.CENTER,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            width=150,
            height=150,
            bgcolor=ft.Colors.WHITE10,
            border_radius=20,
            padding=20,
            on_click=lambda e: start_level(e, level_num),
            ink=True,
            border=ft.border.all(1, ft.Colors.WHITE24)
        )

    grid = ft.Row(
        wrap=True,
        spacing=20,
        run_spacing=20,
        alignment=ft.MainAxisAlignment.CENTER,
        controls=[level_card(i) for i in range(1, 7)]
    )


    def pick_files_result(e: ft.FilePickerResultEvent):
        if e.files:
            file_path = e.files[0].path
            
            def process_import(target_level, level_dlg):
                try:
                    from openpyxl import load_workbook
                    from database import get_db_connection
                    
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    
                    wb = load_workbook(file_path, read_only=True)
                    ws = wb.active
                    
                    count = 0
                    for row in ws.iter_rows(values_only=True):
                        try:
                            # Convert to strings, skip None
                            cells = [str(c).strip() if c is not None else "" for c in row]
                            
                            # Skip header heuristic
                            if cells and ("english" in cells[0].lower() or "level" in cells[0].lower() or "word" in cells[0].lower()):
                                continue
                                
                            en, ar = "", ""
                            if len(cells) == 2:
                                en, ar = cells[0], cells[1]
                            elif len(cells) >= 3:
                                en, ar = cells[1], cells[2]
                                
                            if en and ar:
                                cursor.execute('INSERT INTO words (level, english_word, arabic_word) VALUES (?, ?, ?)', (target_level, en, ar))
                                count += 1
                        except Exception as row_ex:
                            log(f"Skipping row {row}: {row_ex}")
                                
                    wb.close()
                    conn.commit()
                    conn.close()
                    
                    page.close(level_dlg)
                    page.snack_bar = ft.SnackBar(ft.Text(f"Successfully imported {count} words to Level {target_level}!"))
                    page.snack_bar.open = True
                    page.update()
                    
                except Exception as ex:
                    page.snack_bar = ft.SnackBar(ft.Text(f"Error importing: {ex}"))
                    page.snack_bar.open = True
                    page.update()

            # Level selection dialog
            def select_level(e, lvl, level_dlg):
                process_import(lvl, level_dlg)

            level_buttons = [
                ft.ElevatedButton(f"Level {i}", on_click=lambda e, i=i: select_level(e, i, level_dialog)) 
                for i in range(1, 7)
            ]

            level_dialog = ft.AlertDialog(
                title=ft.Text("Select Target Level"),
                content=ft.Column([ft.Text("Which level are these words for?")], height=50),
                actions=level_buttons,
                actions_alignment=ft.MainAxisAlignment.CENTER,
                modal=True
            )
            
            page.open(level_dialog)
            page.update()

    def import_csv_dialog(e):
        file_picker.pick_files(allow_multiple=False, allowed_extensions=["xlsx"])

    file_picker = ft.FilePicker(on_result=pick_files_result)
    
    # Check if a FilePicker is already in page.overlay to avoid "Unknown control" or duplicates
    existing_fp = None
    for c in page.overlay:
        if isinstance(c, ft.FilePicker):
            existing_fp = c
            break
            
    if not existing_fp:
        page.overlay.append(file_picker)
    else:
        file_picker = existing_fp
        file_picker.on_result = pick_files_result

    return ft.View(
        route="/dashboard",
        controls=[
            ft.Container(
                content=ft.Column(
                    [
                        ft.Row(
                            [
                                ft.Column([
                                    ft.Text(f"Welcome, {user['username']}", size=28, weight=ft.FontWeight.BOLD, color="white"),
                                    ft.Text(f"Total Score: {user['score']}", size=16, color="greenAccent"),
                                ]),
                                ft.Row([
                                    ft.PopupMenuButton(
                                        icon=ft.Icons.SETTINGS,
                                        icon_color="white",
                                        tooltip="Settings",
                                        items=[
                                            ft.PopupMenuItem(content=ft.Text("Import Words (XLSX)"), icon=ft.Icons.UPLOAD_FILE, on_click=import_csv_dialog),
                                            ft.PopupMenuItem(content=ft.Text("Manage Words"), icon=ft.Icons.LIST_ALT, on_click=lambda e: page.go("/words")),
                                        ]
                                    ),
                                    ft.IconButton(ft.Icons.LOGOUT, icon_color="redAccent", on_click=lambda e: [(__import__('database').clear_last_user()), (__import__('session_utils').clear_session()), page.go("/")]),
                                ])
                            ],
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                        ),
                        ft.Container(height=40),
                        ft.Text("Select a Level", size=24, color="white", weight=ft.FontWeight.W_500),
                        ft.Container(height=20),
                        grid,
                        ft.Container(height=30),
                        ft.Container(
                            content=ft.Row(
                                [
                                    ft.Icon(ft.Icons.WARNING_AMBER_ROUNDED, color="white70", size=30),
                                    ft.Column(
                                        [
                                            ft.Text("Difficult Words", size=18, weight=ft.FontWeight.BOLD, color="white"),
                                            ft.Text(f"{len(get_difficult_words(username))} words to review", size=13, color="white70"),
                                        ],
                                        spacing=2,
                                    ),
                                    ft.Icon(ft.Icons.ARROW_FORWARD_IOS, color="white54", size=18),
                                ],
                                alignment=ft.MainAxisAlignment.START,
                                spacing=15,
                            ),
                            bgcolor=ft.Colors.BLUE_GREY_900,
                            border_radius=15,
                            padding=ft.Padding(left=20, right=20, top=15, bottom=15),
                            on_click=lambda e: page.go("/difficult"),
                            ink=True,
                        ),
                    ],
                    horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                padding=40,
                expand=True,
                gradient=ft.LinearGradient(
                    begin=ft.Alignment(0, -1),
                    end=ft.Alignment(0, 1),
                    colors=[ft.Colors.BLUE_GREY_900, ft.Colors.BLACK],
                )
            )
        ],
        padding=0
    )
